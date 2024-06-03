# -*- coding: cp1251 -*-
import tkinter as tk
from tkinter import ttk, messagebox
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

def make_request_with_retries(url, retries=5, delay=100):
    for attempt in range(retries):
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException as e:
            print(f"Ошибка подключения: {e} - Попытка {attempt + 1} из {retries}")
            if attempt == retries - 1:
                return None
            time.sleep(delay)
    return None

def collect_links(base_url, page_limit):
    links = set()
    page_number = 1
    while page_number <= page_limit:
        url = f"{base_url}?PAGEN_1={page_number}"
        response = make_request_with_retries(url)
        if response is None:
            break

        soup = BeautifulSoup(response.content, 'html.parser')
        news_items = soup.find_all('div', class_='info')
        
        if not news_items:
            break
        
        for item in news_items:
            link = item.find('a', class_='name')['href']
            full_link = f"https://cchgeu.ru{link}"
            links.add(full_link)
        
        page_number += 1
    
    return links

def collect_news_data(links):
    news_data = set()
    for link in links:
        response = make_request_with_retries(link)
        if response is None:
            continue

        soup = BeautifulSoup(response.content, 'html.parser')
        title_tag = soup.find('h1')
        title = title_tag.get_text(strip=True)
        
        date_tag = soup.find('span', dir='ltr')
        date_str = date_tag.get_text(strip=True).strip()
        for rus, eng in months.items():
            date_str = date_str.replace(rus, eng)
        date = datetime.strptime(date_str, '%d %B %Y')
        tags = [tag.get_text(strip=True) for tag in soup.select('a[href^="/press/news/?tag="]')]
        tags_str = ", ".join(tags) if tags else "Нет тегов"
        
        news_data.add((title, link, date, tags_str))
    
    return news_data

def save_data_to_sheet(ws, data, old_data):
    ws.delete_rows(2, ws.max_row)
    for title, link, date, tags_str in data:
        status = "Старая"
        if (title, link, date, tags_str) not in old_data:
            status = "Новая"
        ws.append([title, link, date, tags_str, status])

months = {
    "января": "January", "февраля": "February", "марта": "March",
    "апреля": "April", "мая": "May", "июня": "June",
    "июля": "July", "августа": "August", "сентября": "September",
    "октября": "October", "ноября": "November", "декабря": "December"
}

sections = {
    "Официальные новости": "https://cchgeu.ru/press/news/official/",
    "Объявления": "https://cchgeu.ru/press/news/ad/",
    "СМИ о нас": "https://cchgeu.ru/press/news/digest/",
    "Выступления, доклады, интервью": "https://cchgeu.ru/press/news/interview/"
}

file_path = "NEWS.xlsx"
all_news_file_path = "ALL_NEWS.xlsx"

class NewsParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("News Parser")

        self.root.rowconfigure(4, weight=1)

        self.section_var = tk.StringVar()
        self.page_number_var = tk.IntVar(value=3)

        self.create_widgets()

    def create_widgets(self):
        # Section selection
        section_label = tk.Label(self.root, text="Выберите раздел для парсинга:")
        section_label.grid(row=0, column=1, columnspan=2, padx=10, pady=5, sticky="e")
        
        self.section_combobox = ttk.Combobox(self.root, textvariable=self.section_var, values=list(sections.keys()))
        self.section_combobox.grid(row=0, column=3, columnspan=2, padx=10, pady=5, sticky="nsew")

        # Page number entry
        page_label = tk.Label(self.root, text="Количество страниц:")
        page_label.grid(row=1, column=1, columnspan=2, padx=10, pady=5, sticky="e")
        
        self.page_entry = tk.Entry(self.root, textvariable=self.page_number_var)
        self.page_entry.grid(row=1, column=3,columnspan=2, padx=10, pady=5, sticky="nsew")

        # Parse button
        parse_button = tk.Button(self.root, text="Парсить", command=self.parse_news)
        parse_button.grid(row=0, column=0, columnspan=1, pady=10, sticky="nsew")

        # Show button
        show_button = tk.Button(self.root, text="Показать новости", command=self.show_news)
        show_button.grid(row=1, column=0, columnspan=1, pady=10, sticky="nsew")
        
        self.search_entry = tk.Entry(self.root)
        self.search_entry.grid(row=2, column=0, sticky="nsew")

        self.search_button = tk.Button(self.root, text="Поиск", command=self.search_tags)
        self.search_button.grid(row=2, column=1, sticky="nsew")

        self.open_button_NEWS = tk.Button(self.root, text="Открыть файл с последней операцией", command=self.open_NEWS)
        self.open_button_NEWS.grid(row=2, column=3, sticky="nsew")

        self.open_button_ALL_NEWS = tk.Button(self.root, text="Открыть файл со всеми новостями", command=self.open_ALL_NEWS)
        self.open_button_ALL_NEWS.grid(row=2, column=4, sticky="nsew")

        # Tables
        all_news_label = tk.Label(self.root, text="Все новости:")
        all_news_label.grid(row=3, column=2, columnspan=2, pady=10, sticky="nsew")

        self.all_news_table = ttk.Treeview(self.root, columns=("Название", "Дата", "Теги"), show="headings")
        self.all_news_table.heading("Название", text="Название")
        self.all_news_table.heading("Дата", text="Дата")
        self.all_news_table.heading("Теги", text="Теги")
        self.all_news_table.grid(row=4, column=2, columnspan=3, padx=10, pady=10, sticky="nsew")

        latest_news_label = tk.Label(self.root, text="Новости последней операции:")
        latest_news_label.grid(row=3, column=0, columnspan=2, pady=10,  sticky="nsew")

        self.latest_news_table = ttk.Treeview(self.root, columns=("Название", "Дата", "Теги", "Статус"), show="headings")
        self.latest_news_table.heading("Название", text="Название")
        self.latest_news_table.heading("Дата", text="Дата")
        self.latest_news_table.heading("Теги", text="Теги")
        self.latest_news_table.heading("Статус", text="Статус")
        self.latest_news_table.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

    def parse_news(self):
        section = self.section_var.get()
        page_number = self.page_number_var.get()

        if not section or page_number < 1:
            messagebox.showerror("Ошибка", "Выберите раздел и укажите количество страниц больше 0")
            return

        base_url = sections[section]
        sheet_title = section

        old_news_data = set()
        all_news_data = set()

        if os.path.exists(file_path):
            wb_old = load_workbook(file_path)
            if sheet_title in wb_old.sheetnames:
                ws_old_news = wb_old[sheet_title]
                for row in ws_old_news.iter_rows(min_row=2, values_only=True):
                    old_news_data.add((row[0], row[1], row[2], row[3]))
            else:
                ws_old_news = wb_old.create_sheet(sheet_title)
        else:
            wb_old = Workbook()
            ws_old_news = wb_old.active
            ws_old_news.title = sheet_title

        if ws_old_news.max_row == 1:
            ws_old_news.append(["Название новости", "Ссылка", "Дата публикации", "Теги", "Статус"])

        if os.path.exists(all_news_file_path):
            wb_all_old = load_workbook(all_news_file_path)
            if sheet_title in wb_all_old.sheetnames:
                ws_all_old_news = wb_all_old[sheet_title]
                for row in ws_all_old_news.iter_rows(min_row=2, values_only=True):
                    all_news_data.add((row[0], row[1], row[2], row[3]))
            else:
                ws_all_old_news = wb_all_old.create_sheet(sheet_title)
        else:
            wb_all_old = Workbook()
            ws_all_old_news = wb_all_old.active
            ws_all_old_news.title = sheet_title

        if ws_all_old_news.max_row == 1:
            ws_all_old_news.append(["Название новости", "Ссылка", "Дата публикации", "Теги"])

        news_links = collect_links(base_url, page_number)
        news_data = collect_news_data(news_links)

        save_data_to_sheet(ws_old_news, news_data, old_news_data)
        wb_old.save(file_path)
        print(f"Новости сохранены в файл {file_path}")

        for title, link, date, tags_str in news_data:
            if (title, link, date, tags_str) not in all_news_data:
                ws_all_old_news.append([title, link, date, tags_str])
                all_news_data.add((title, link, date, tags_str))
        wb_all_old.save(all_news_file_path)
        print(f"Все новости сохранены в файл {all_news_file_path}")

        self.populate_all_news_table(section)
        self.populate_latest_news_table(section)
        
    def search_tags(self):
        search_tag = self.search_entry.get()
        self.all_news_table.selection_remove(self.all_news_table.selection())
    
        for row in self.all_news_table.get_children():
            tags = self.all_news_table.item(row)["values"][2] 
            if search_tag in tags:
                self.all_news_table.selection_add(row)
        else:
            messagebox.showinfo("Поиск", f"Тег '{search_tag}' не найден.")
            
    def open_NEWS(self):
        os.startfile(file_path)

    def open_ALL_NEWS(self):
        os.startfile(all_news_file_path)

    def show_news(self):
        section = self.section_var.get()
        if not section:
            messagebox.showerror("Ошибка", "Выберите раздел для отображения новостей")
            return

        self.populate_all_news_table(section)
        self.populate_latest_news_table(section)

    def populate_all_news_table(self, section):
        for i in self.all_news_table.get_children():
            self.all_news_table.delete(i)

        if os.path.exists(all_news_file_path):
            wb = load_workbook(all_news_file_path)
            if section in wb.sheetnames:
                ws = wb[section]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    self.all_news_table.insert("", "end", values=(row[0], row[2], row[3]))

    def populate_latest_news_table(self, section):
        for i in self.latest_news_table.get_children():
            self.latest_news_table.delete(i)

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if section in wb.sheetnames:
                ws = wb[section]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    self.latest_news_table.insert("", "end", values=(row[0], row[2], row[3], row[4]))

if __name__ == "__main__":
    root = tk.Tk()
    app = NewsParserApp(root)
    root.mainloop()